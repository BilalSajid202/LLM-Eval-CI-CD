"""Export evaluation reports to multi-sheet Excel workbooks with charts."""

from __future__ import annotations

from pathlib import Path

from llm_eval.models.types import EvaluationReport, QuestionResult


def _require_openpyxl():
    try:
        import openpyxl
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        return openpyxl, BarChart, Reference, Alignment, Font, PatternFill, get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install 'llm-eval[reporting]'"
        ) from exc


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(cell.value or "")) for cell in col), default=min_width)
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def _write_summary_sheet(ws, report: EvaluationReport) -> None:
    _, _, _, Alignment, Font, PatternFill, _ = _require_openpyxl()

    header_font = Font(bold=True, size=14)
    ws["A1"] = "LLM Evaluation Report — Summary"
    ws["A1"].font = header_font

    run = report.run
    rows = [
        ("Run ID", str(run.run_id)),
        ("Status", run.status.value.upper()),
        ("Scope", run.scope),
        ("Git SHA", run.git_sha[:12]),
        ("Branch", run.git_branch),
        ("Model", run.model_version),
        ("Started", str(run.started_at)),
        ("Finished", str(run.finished_at or "")),
        ("Questions", report.question_count),
        ("Overall Rank", f"{report.overall_rank}/{report.total_runs_compared}"
         if report.overall_rank else "N/A"),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    if run.metrics:
        m = run.metrics
        ws.cell(row=15, column=1, value="Aggregate Metrics").font = Font(bold=True, size=12)
        metrics_rows = [
            ("Hallucination Rate", f"{m.hallucination_rate:.2%}"),
            ("Answer Relevancy", f"{m.answer_relevancy:.3f}"),
            ("Faithfulness", f"{m.faithfulness:.3f}"),
            ("Context Recall", f"{m.context_recall:.3f}"),
            ("Accuracy", f"{m.accuracy:.3f}"),
            ("Precision", f"{m.precision:.3f}"),
            ("Recall", f"{m.recall:.3f}"),
            ("F1 Score", f"{m.f1_score:.3f}"),
            ("Injection Resistance", f"{m.prompt_injection_resistance:.1%}"),
            ("Jailbreak Resistance", f"{m.jailbreak_resistance:.1%}"),
            ("p50 Latency (ms)", f"{m.p50_latency_ms:.0f}"),
            ("p95 Latency (ms)", f"{m.p95_latency_ms:.0f}"),
            ("Cost/Query (USD)", f"${m.cost_per_query_usd:.6f}"),
            ("Errors", str(m.error_count)),
        ]
        for i, (label, value) in enumerate(metrics_rows, start=16):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

    row = 32
    ws.cell(row=row, column=1, value="Gate Results").font = Font(bold=True, size=12)
    row += 1
    for g in run.gate_results:
        ws.cell(row=row, column=1, value=g.metric)
        ws.cell(row=row, column=2, value=g.status.value.upper())
        ws.cell(row=row, column=3, value=g.message)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Key Insights").font = Font(bold=True, size=12)
    row += 1
    for insight in report.insights:
        ws.cell(row=row, column=1, value=f"• {insight}")
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Recommendations").font = Font(bold=True, size=12)
    row += 1
    for rec in report.recommendations:
        ws.cell(row=row, column=1, value=f"• {rec}")
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    _auto_width(ws)


def _write_raw_results_sheet(ws, results: list[QuestionResult]) -> None:
    _, _, _, _, Font, _, _ = _require_openpyxl()

    headers = [
        "Question ID", "Category", "Tags", "Question", "Expected Answer", "Answer",
        "Accuracy", "Relevancy", "Precision", "Recall", "F1", "Faithfulness",
        "Context Recall", "Hallucination", "Injection Score", "Jailbreak Score",
        "Pass", "Latency (ms)", "Cost (USD)", "Error",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, r in enumerate(results, start=2):
        tags = ", ".join(r.metadata.get("tags", []))
        ws.cell(row=row_idx, column=1, value=r.question_id)
        ws.cell(row=row_idx, column=2, value=r.category)
        ws.cell(row=row_idx, column=3, value=tags)
        ws.cell(row=row_idx, column=4, value=r.question[:200])
        ws.cell(row=row_idx, column=5, value=r.expected_answer[:200])
        ws.cell(row=row_idx, column=6, value=r.answer[:300])
        ws.cell(row=row_idx, column=7, value=round(r.scores.get("accuracy", 0), 3))
        ws.cell(row=row_idx, column=8, value=round(r.scores.get("answer_relevancy", 0), 3))
        ws.cell(row=row_idx, column=9, value=round(r.scores.get("precision", 0), 3))
        ws.cell(row=row_idx, column=10, value=round(r.scores.get("recall", 0), 3))
        ws.cell(row=row_idx, column=11, value=round(r.scores.get("f1_score", 0), 3))
        ws.cell(row=row_idx, column=12, value=round(r.scores.get("faithfulness", 0), 3))
        ws.cell(row=row_idx, column=13, value=round(r.scores.get("context_recall", 0), 3))
        hall = r.scores.get("hallucination_verdict", 0)
        ws.cell(row=row_idx, column=14, value="Yes" if hall >= 0.5 else "No")
        inj = r.scores.get("prompt_injection_score")
        jb = r.scores.get("jailbreak_score")
        ws.cell(row=row_idx, column=15, value=round(inj, 3) if inj is not None else "N/A")
        ws.cell(row=row_idx, column=16, value=round(jb, 3) if jb is not None else "N/A")
        ws.cell(row=row_idx, column=17, value="PASS" if r.metadata.get("accuracy_pass") else "FAIL")
        ws.cell(row=row_idx, column=18, value=round(r.latency_ms, 0))
        ws.cell(row=row_idx, column=19, value=round(r.cost_usd, 6))
        ws.cell(row=row_idx, column=20, value=r.error or "")

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_metric_comparison_sheet(ws, report: EvaluationReport) -> None:
    _, _, _, _, Font, _, _ = _require_openpyxl()

    headers = ["Metric", "Current", "Baseline", "Delta", "Delta %", "Trend", "Direction"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    for row_idx, c in enumerate(report.comparisons, start=2):
        ws.cell(row=row_idx, column=1, value=c.metric)
        ws.cell(row=row_idx, column=2, value=round(c.current, 4))
        ws.cell(row=row_idx, column=3, value=round(c.baseline, 4))
        ws.cell(row=row_idx, column=4, value=round(c.delta, 4))
        ws.cell(row=row_idx, column=5, value=f"{c.delta_pct:.1f}%")
        ws.cell(row=row_idx, column=6, value=c.trend.upper())
        ws.cell(row=row_idx, column=7, value=c.direction)

    _auto_width(ws)


def _write_category_sheet(ws, report: EvaluationReport) -> None:
    _, _, _, _, Font, _, _ = _require_openpyxl()

    headers = ["Category", "Questions", "Accuracy", "Relevancy", "Hallucination Rate", "Pass Rate"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    for row_idx, cat in enumerate(report.category_breakdown, start=2):
        ws.cell(row=row_idx, column=1, value=cat.category)
        ws.cell(row=row_idx, column=2, value=cat.question_count)
        ws.cell(row=row_idx, column=3, value=round(cat.accuracy, 3))
        ws.cell(row=row_idx, column=4, value=round(cat.answer_relevancy, 3))
        ws.cell(row=row_idx, column=5, value=round(cat.hallucination_rate, 3))
        ws.cell(row=row_idx, column=6, value=f"{cat.pass_rate:.1%}")

    _auto_width(ws)


def _write_charts_sheet(ws, report: EvaluationReport) -> None:
    openpyxl, BarChart, Reference, _, Font, _, get_column_letter = _require_openpyxl()

    ws.cell(row=1, column=1, value="Metric Visualizations").font = Font(bold=True, size=14)

    if report.run.metrics:
        m = report.run.metrics
        chart_metrics = [
            ("Accuracy", m.accuracy),
            ("Relevancy", m.answer_relevancy),
            ("Faithfulness", m.faithfulness),
            ("F1 Score", m.f1_score),
            ("Precision", m.precision),
            ("Recall", m.recall),
            ("Injection Res.", m.prompt_injection_resistance),
            ("Jailbreak Res.", m.jailbreak_resistance),
        ]
        ws.cell(row=3, column=1, value="Metric").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Score").font = Font(bold=True)
        for i, (name, val) in enumerate(chart_metrics, start=4):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=round(val, 3))

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Quality Metrics (higher is better)"
        chart1.y_axis.title = "Score"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(chart_metrics))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(chart_metrics))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.width = 18
        chart1.height = 10
        ws.add_chart(chart1, "D3")

    if report.comparisons:
        start_row = 15
        ws.cell(row=start_row, column=1, value="Metric").font = Font(bold=True)
        ws.cell(row=start_row, column=2, value="Current").font = Font(bold=True)
        ws.cell(row=start_row, column=3, value="Baseline").font = Font(bold=True)
        for i, c in enumerate(report.comparisons[:8], start=start_row + 1):
            ws.cell(row=i, column=1, value=c.metric)
            ws.cell(row=i, column=2, value=round(c.current, 4))
            ws.cell(row=i, column=3, value=round(c.baseline, 4))

        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Current vs Baseline"
        chart2.y_axis.title = "Value"
        data = Reference(
            ws, min_col=2, min_row=start_row, max_col=3,
            max_row=start_row + min(8, len(report.comparisons)),
        )
        cats = Reference(
            ws, min_col=1, min_row=start_row + 1,
            max_row=start_row + min(8, len(report.comparisons)),
        )
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.width = 18
        chart2.height = 10
        ws.add_chart(chart2, "D15")

    _auto_width(ws)


def export_excel(report: EvaluationReport, results: list[QuestionResult], path: Path) -> None:
    openpyxl, _, _, _, _, _, _ = _require_openpyxl()

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, report)

    ws_raw = wb.create_sheet("Raw Results")
    _write_raw_results_sheet(ws_raw, results)

    ws_compare = wb.create_sheet("Metric Comparisons")
    _write_metric_comparison_sheet(ws_compare, report)

    ws_category = wb.create_sheet("Category Breakdown")
    _write_category_sheet(ws_category, report)

    ws_charts = wb.create_sheet("Visualizations")
    _write_charts_sheet(ws_charts, report)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
